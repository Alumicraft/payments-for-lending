"""
Dealer Import from Contact List XLSX

Imports dealers from the "Contact List 020526.xlsx" spreadsheet into ERPNext,
creating Customers, Addresses, Contacts, Suppliers (factories), and
Factory Assignments.

Usage (browser console):
    frappe.call({method: "dcr.api.dealer_import.import_dealers_from_xlsx"})
    frappe.call({method: "dcr.api.dealer_import.import_dealers_from_xlsx", args: {dry_run: 1}})
    frappe.call({method: "dcr.api.dealer_import.import_dealers_from_xlsx", args: {limit: 5}})
"""

import re

import frappe
from frappe import _
from frappe.utils import today

import openpyxl


FACTORY_MAP = {
    "Dur": "Durango Homes",
    "Wst": "Skyline Homes",
    "FltW": "Fleetwood Homes",
    "Chp": "Champion Home Builders",
}

# Column indices (1-based) in the Master sheet
COL_DUR = 2    # B
COL_WST = 3    # C
COL_FLTW = 4   # D
COL_CHP = 5    # E
COL_FIRST = 6  # F
COL_LAST = 7   # G
COL_DBA = 8    # H
COL_LEGAL = 9  # I
COL_BIZ_PHONE = 10   # J
COL_MOBILE = 11      # K
COL_OTHER_PHONE = 12  # L
COL_EMAIL = 13        # M
COL_EMAIL2 = 14       # N
COL_ADDRESS = 15      # O


def parse_address(raw):
    """Parse a raw address string into components.

    Returns dict with address_line1, city, state, pincode or None if unparseable.
    """
    if not raw:
        return None

    addr = str(raw).strip()

    # Skip note addresses
    if any(x in addr.upper() for x in ["SEND MCO", "PICK UP"]):
        return None

    # Strip prefixes like (Residential), (R) at start
    addr = re.sub(r"^\(.*?\)\s*", "", addr)
    # Strip trailing (R)
    addr = re.sub(r"\s*\(R\)\s*$", "", addr)

    # Match: street, city, STATE ZIP (two commas)
    match = re.match(r"^(.+),\s*([^,]+),\s*([A-Z]{2})\s+(\d{5})\s*$", addr)
    if match:
        return {
            "address_line1": match.group(1).strip(),
            "city": match.group(2).strip(),
            "state": match.group(3),
            "pincode": match.group(4),
        }

    # Fallback: street city, STATE ZIP (one comma — no comma between street and city)
    match = re.match(r"^(.+),\s*([A-Z]{2})\s+(\d{5})\s*$", addr)
    if match:
        street_city = match.group(1).strip()
        parts = street_city.rsplit(" ", 1)
        if len(parts) == 2:
            return {
                "address_line1": parts[0].strip(),
                "city": parts[1].strip().title(),
                "state": match.group(2),
                "pincode": match.group(3),
            }

    return None


def clean_phone(val):
    """Convert phone value to clean string."""
    if val is None:
        return None
    s = str(val).replace("\xa0", "").strip()
    if not s:
        return None
    # If it looks like a float from Excel (e.g. 7604711234.0), strip the decimal
    if s.endswith(".0"):
        s = s[:-2]
    # Strip extensions: "Ext. 2", "ext 2", "x2", etc.
    s = re.sub(r'\s*(ext\.?|x)\s*\d+$', '', s, flags=re.IGNORECASE).strip()
    return s or None


def clean_customer_name(dba):
    """Strip commas from DBA to avoid ERPNext naming issues."""
    if not dba:
        return None
    return str(dba).replace(",", "").strip()


def cell_val(ws, row, col):
    """Get cell value, returning None for empty cells."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        return val if val else None
    return val


def is_factory_flag(val):
    """Check if cell value indicates a factory assignment (X or x)."""
    if val is None:
        return False
    return str(val).strip().upper() == "X"


@frappe.whitelist()
def import_dealers_from_xlsx(dry_run=False, limit=0):
    """Import dealers from the Contact List XLSX into ERPNext.

    Args:
        dry_run: If True, do not create any records, just report what would happen.
        limit: If > 0, only process this many unique companies.

    Returns:
        dict with counts of created/skipped records and any errors.
    """
    dry_run = frappe.utils.cint(dry_run)
    limit = frappe.utils.cint(limit)

    site = frappe.local.site
    file_path = f"/home/frappe/frappe-bench/sites/{site}/public/files/Contact List 020526.xlsx"

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb["Master"]

    # --- Read all rows into structured data ---
    rows = []
    for row_num in range(4, ws.max_row + 1):
        dba = cell_val(ws, row_num, COL_DBA)
        first_name = cell_val(ws, row_num, COL_FIRST)
        if not dba and not first_name:
            continue  # skip empty rows

        factories = []
        if is_factory_flag(cell_val(ws, row_num, COL_DUR)):
            factories.append("Dur")
        if is_factory_flag(cell_val(ws, row_num, COL_WST)):
            factories.append("Wst")
        if is_factory_flag(cell_val(ws, row_num, COL_FLTW)):
            factories.append("FltW")
        if is_factory_flag(cell_val(ws, row_num, COL_CHP)):
            factories.append("Chp")

        rows.append({
            "dba": dba,
            "legal_name": cell_val(ws, row_num, COL_LEGAL),
            "first_name": first_name,
            "last_name": cell_val(ws, row_num, COL_LAST),
            "biz_phone": clean_phone(cell_val(ws, row_num, COL_BIZ_PHONE)),
            "mobile": clean_phone(cell_val(ws, row_num, COL_MOBILE)),
            "other_phone": clean_phone(cell_val(ws, row_num, COL_OTHER_PHONE)),
            "email": cell_val(ws, row_num, COL_EMAIL),
            "email2": cell_val(ws, row_num, COL_EMAIL2),
            "address": cell_val(ws, row_num, COL_ADDRESS),
            "factories": factories,
        })

    wb.close()

    # --- Group by DBA (company) ---
    companies = {}
    for row in rows:
        cname = clean_customer_name(row["dba"])
        if not cname:
            continue
        if cname not in companies:
            companies[cname] = {
                "customer_name": cname,
                "legal_name": row.get("legal_name"),
                "contacts": [],
                "factories": set(),
            }
        companies[cname]["contacts"].append(row)
        companies[cname]["factories"].update(row["factories"])

    if limit:
        company_names = list(companies.keys())[:limit]
        companies = {k: companies[k] for k in company_names}

    # --- Counters ---
    stats = {
        "created_suppliers": 0,
        "created_customers": 0,
        "created_addresses": 0,
        "created_contacts": 0,
        "created_assignments": 0,
        "skipped": 0,
        "errors": [],
    }

    # --- Step 0: Ensure Supplier Group "Factory" and Factory Suppliers ---
    if not dry_run:
        _ensure_supplier_group(stats)
        _ensure_factory_suppliers(stats)
    else:
        if not frappe.db.exists("Supplier Group", "Factory"):
            stats["created_suppliers"] += 1  # the group
        for factory_name in FACTORY_MAP.values():
            if not frappe.db.exists("Supplier", factory_name):
                stats["created_suppliers"] += 1

    # --- Process each company ---
    processed = 0
    for cname, company in companies.items():
        try:
            first_contact = company["contacts"][0]

            # --- Step 1: Customer ---
            if frappe.db.exists("Customer", cname):
                stats["skipped"] += 1
            elif not dry_run:
                _create_customer(cname)
                stats["created_customers"] += 1
            else:
                stats["created_customers"] += 1

            # --- Step 2: Address ---
            addr_data = None
            for contact in company["contacts"]:
                addr_data = parse_address(contact.get("address"))
                if addr_data:
                    break

            if addr_data:
                addr_exists = frappe.db.exists("Address", {
                    "address_title": cname,
                    "address_type": "Billing",
                })
                if not addr_exists:
                    if not dry_run:
                        _create_address(cname, addr_data, first_contact)
                        stats["created_addresses"] += 1
                    else:
                        stats["created_addresses"] += 1

            # --- Step 3: Contacts ---
            for idx, contact in enumerate(company["contacts"]):
                is_primary = idx == 0
                first_name = contact.get("first_name") or ""
                last_name = contact.get("last_name") or ""

                if not first_name and not last_name:
                    continue

                contact_exists = _contact_exists(first_name, last_name, cname)
                if not contact_exists:
                    if not dry_run:
                        _create_contact(cname, contact, is_primary)
                        stats["created_contacts"] += 1
                    else:
                        stats["created_contacts"] += 1

            # --- Step 4: Factory Assignments ---
            for factory_key in company["factories"]:
                supplier_name = FACTORY_MAP.get(factory_key)
                if not supplier_name:
                    continue

                fa_exists = frappe.db.exists("Factory Assignment", {
                    "customer": cname,
                    "factory": supplier_name,
                })
                if not fa_exists:
                    if not dry_run:
                        _create_factory_assignment(cname, supplier_name)
                        stats["created_assignments"] += 1
                    else:
                        stats["created_assignments"] += 1

            processed += 1

            # Batch commit every 10 customers
            if not dry_run and processed % 10 == 0:
                frappe.db.commit()

        except Exception as e:
            stats["errors"].append(f"{cname}: {str(e)}")
            frappe.log_error(f"Dealer import error for {cname}", str(e))

    # Final commit
    if not dry_run:
        frappe.db.commit()

    stats["dry_run"] = bool(dry_run)
    stats["total_companies"] = len(companies)
    return stats


def _ensure_supplier_group(stats):
    """Create Supplier Group 'Factory' if it doesn't exist."""
    if not frappe.db.exists("Supplier Group", "Factory"):
        doc = frappe.get_doc({
            "doctype": "Supplier Group",
            "supplier_group_name": "Factory",
        })
        doc.insert(ignore_permissions=True)
        stats["created_suppliers"] += 1


def _ensure_factory_suppliers(stats):
    """Create the four factory Supplier records if they don't exist."""
    for factory_name in FACTORY_MAP.values():
        if not frappe.db.exists("Supplier", factory_name):
            doc = frappe.get_doc({
                "doctype": "Supplier",
                "supplier_name": factory_name,
                "supplier_group": "Factory",
                "supplier_type": "Company",
                "country": "United States",
            })
            doc.insert(ignore_permissions=True)
            stats["created_suppliers"] += 1


def _create_customer(customer_name):
    """Create a Customer record."""
    doc = frappe.get_doc({
        "doctype": "Customer",
        "customer_name": customer_name,
        "customer_type": "Company",
        "customer_group": "Dealer",
        "territory": "United States",
        "default_currency": "USD",
    })
    doc.insert(ignore_permissions=True)


def _create_address(customer_name, addr_data, first_contact):
    """Create an Address linked to a Customer."""
    phone = first_contact.get("biz_phone") or first_contact.get("mobile")
    email = first_contact.get("email")

    doc = frappe.get_doc({
        "doctype": "Address",
        "address_title": customer_name,
        "address_type": "Billing",
        "address_line1": addr_data["address_line1"],
        "city": addr_data["city"],
        "state": addr_data["state"],
        "pincode": addr_data["pincode"],
        "country": "United States",
        "is_primary_address": 1,
        "is_shipping_address": 1,
        "phone": phone,
        "email_id": email,
        "links": [{
            "link_doctype": "Customer",
            "link_name": customer_name,
        }],
    })
    doc.insert(ignore_permissions=True)


def _contact_exists(first_name, last_name, customer_name):
    """Check if a Contact already exists for this person linked to this Customer."""
    contacts = frappe.db.get_all("Contact", filters={
        "first_name": first_name,
        "last_name": last_name or "",
    }, fields=["name"])

    for c in contacts:
        links = frappe.db.get_all("Dynamic Link", filters={
            "parent": c.name,
            "parenttype": "Contact",
            "link_doctype": "Customer",
            "link_name": customer_name,
        })
        if links:
            return True

    return False


def _create_contact(customer_name, contact_data, is_primary):
    """Create a Contact linked to a Customer."""
    doc = frappe.get_doc({
        "doctype": "Contact",
        "first_name": contact_data.get("first_name") or "",
        "last_name": contact_data.get("last_name") or "",
        "is_primary_contact": 1 if is_primary else 0,
        "is_billing_contact": 1 if is_primary else 0,
        "links": [{
            "link_doctype": "Customer",
            "link_name": customer_name,
        }],
    })

    # Email addresses
    if contact_data.get("email"):
        doc.append("email_ids", {
            "email_id": contact_data["email"],
            "is_primary": 1,
        })
    if contact_data.get("email2"):
        doc.append("email_ids", {
            "email_id": contact_data["email2"],
            "is_primary": 0,
        })

    # Phone numbers
    if contact_data.get("mobile"):
        doc.append("phone_nos", {
            "phone": contact_data["mobile"],
            "is_primary_mobile_no": 1,
        })
    if contact_data.get("biz_phone"):
        doc.append("phone_nos", {
            "phone": contact_data["biz_phone"],
            "is_primary_phone": 1,
        })
    if contact_data.get("other_phone"):
        doc.append("phone_nos", {
            "phone": contact_data["other_phone"],
        })

    doc.insert(ignore_permissions=True)


def _create_factory_assignment(customer_name, supplier_name):
    """Create a Factory Assignment linking a Customer to a factory Supplier."""
    doc = frappe.get_doc({
        "doctype": "Factory Assignment",
        "customer": customer_name,
        "factory": supplier_name,
        "assignment_date": today(),
        "active": 1,
        "retailer_application_status": "Approved",
    })
    doc.insert(ignore_permissions=True)
    doc.submit()
